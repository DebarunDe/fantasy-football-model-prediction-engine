import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from adp_comparison import create_adp_comparison_sheet
import numpy as np
import csv

def rank_players(df, points_col='weighted_fantasy_points'):
    df = df.copy()
    df['rank'] = df[points_col].rank(ascending=False, method='min')
    df = df.sort_values('rank')
    return df

def export_to_excel(df, filename=None, league_size=12):
    """
    Export the fantasy football big board to Excel with date in filename.
    """
    # Generate filename with current date if not provided
    if filename is None:
        from datetime import datetime
        current_date = datetime.now().strftime("%m%d%y")  # MMDDYY format
        filename = f'fantasy_big_board_{current_date}.xlsx'
    """
    Export the unified big board to Excel with clean, user-friendly formatting.
    Shows ADP comparison first, then unified big board, then position-specific rankings.
    """
    # Select key columns for the unified big board
    columns = [
        'unified_rank', 'player_id', 'team', 'position', 'raw_fantasy_points',
        'unified_big_board_score', 'vor_final'
    ]

    # Only include columns that exist in the dataframe
    export_columns = [col for col in columns if col in df.columns]

    # Add any missing columns with default values
    for col in columns:
        if col not in df.columns:
            if col == 'unified_rank':
                df[col] = df.get('rank', range(1, len(df) + 1))
            elif col in ['unified_big_board_score', 'vor_final']:
                df[col] = 0.0

    # Reorder columns for export
    final_columns = [col for col in columns if col in df.columns]
    export_df = df[final_columns].copy()
    
    # Add DRAFTED column first
    export_df.insert(0, 'DRAFTED', 'NO')

    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Create ADP comparison sheet first (new format)
        create_new_adp_comparison_sheet(writer, df, league_size)
        # Create FantasyPros ADP comparison as second sheet
        create_fantasypros_adp_comparison_sheet(writer, league_size)
        # Write main unified big board
        export_df.to_excel(writer, sheet_name='UNIFIED_BIG_BOARD', index=False)

        # Create position-specific ranking sheets
        create_position_sheets(writer, df)

            # Auto-adjust column widths for unified big board
    worksheet = writer.sheets['UNIFIED_BIG_BOARD']
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add conditional formatting for drafted players
    add_conditional_formatting(worksheet, len(export_df.columns))

def create_position_sheets(writer, df):
    """
    Create position-specific ranking sheets.
    """
    # Position-specific rankings with key metrics
    for pos in ['QB', 'RB', 'WR', 'TE']:
        pos_df = df[df['position'] == pos].copy()
        if not pos_df.empty:
            # Select key columns for position sheets
            pos_columns = ['unified_rank', 'player_id', 'team', 'raw_fantasy_points', 'unified_big_board_score']
            if 'vor_final' in pos_df.columns:
                pos_columns.append('vor_final')
            
            # Only include columns that exist
            pos_export_columns = [col for col in pos_columns if col in pos_df.columns]
            pos_export_df = pos_df[pos_export_columns].copy()
            
            # Add DRAFTED column first
            pos_export_df.insert(0, 'DRAFTED', 'NO')
            
            # Sort by unified rank
            pos_export_df = pos_export_df.sort_values('unified_rank')
            
            pos_export_df.to_excel(writer, sheet_name=f'{pos}_Rankings', index=False)
            
            # Auto-adjust column widths for position sheets
            worksheet = writer.sheets[f'{pos}_Rankings']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add conditional formatting for drafted players
            add_conditional_formatting(worksheet, len(pos_export_df.columns))

def create_new_adp_comparison_sheet(writer, df, league_size=12):
    """
    Create ADP comparison sheet in the new format: ADP, QB, WR, RB, TE, UNIFIED BIG BOARD RANKING, metrics
    """
    # Get ADP comparison data
    adp_comparison_df = create_adp_comparison_sheet(df, league_size)
    
    # Create new format DataFrame
    new_adp_df = pd.DataFrame()
    
    # Add DRAFTED column first
    new_adp_df['DRAFTED'] = 'NO'
    
    # Add ADP column
    new_adp_df['ADP'] = adp_comparison_df['adp']
    
    # Add position columns (empty by default)
    new_adp_df['QB'] = ''
    new_adp_df['WR'] = ''
    new_adp_df['RB'] = ''
    new_adp_df['TE'] = ''
    
    # Fill position columns based on player position
    for idx, row in adp_comparison_df.iterrows():
        position = row['position']
        player_name = row['player_id']
        
        if position == 'QB':
            new_adp_df.at[idx, 'QB'] = player_name
        elif position == 'WR':
            new_adp_df.at[idx, 'WR'] = player_name
        elif position == 'RB':
            new_adp_df.at[idx, 'RB'] = player_name
        elif position == 'TE':
            new_adp_df.at[idx, 'TE'] = player_name
    
    # Add unified big board ranking
    new_adp_df['UNIFIED BIG BOARD RANKING'] = adp_comparison_df['unified_rank']
    
    # Add additional metrics
    if 'raw_fantasy_points' in adp_comparison_df.columns:
        new_adp_df['PROJECTED POINTS'] = adp_comparison_df['raw_fantasy_points'].round(1)
    
    if 'unified_big_board_score' in adp_comparison_df.columns:
        new_adp_df['UNIFIED SCORE'] = adp_comparison_df['unified_big_board_score'].round(3)
    
    if 'value_recommendation' in adp_comparison_df.columns:
        new_adp_df['VALUE RECOMMENDATION'] = adp_comparison_df['value_recommendation']
    
    if 'rank_difference' in adp_comparison_df.columns:
        new_adp_df['RANK DIFFERENCE'] = adp_comparison_df['rank_difference']
    
    # Sort by ADP (handle NaN values)
    new_adp_df = new_adp_df.sort_values('ADP', na_position='last')
    
    # Write to Excel
    new_adp_df.to_excel(writer, sheet_name='ADP_COMPARISON', index=False)
    
    # Get the worksheet for formatting
    worksheet = writer.sheets['ADP_COMPARISON']
    
    # Define color fills
    color_fills = {
        'teal': PatternFill(start_color='00CED1', end_color='00CED1', fill_type='solid'),
        'green': PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid'),
        'light_green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
        'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),  # Neutral - no color
        'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'purple': PatternFill(start_color='800080', end_color='800080', fill_type='solid'),
        'black': PatternFill(start_color='000000', end_color='000000', fill_type='solid')  # For drafted players
    }
    
    # Apply color coding to rows based on value_color (but skip if player is drafted)
    for row_idx, (_, player) in enumerate(adp_comparison_df.iterrows(), start=2):  # Start at 2 to skip header
        # Check if this player is marked as drafted
        drafted_cell = worksheet.cell(row=row_idx, column=1)  # Column A is DRAFTED
        is_drafted = drafted_cell.value == "YES"
        
        if not is_drafted:  # Only apply value colors if not drafted
            color = player.get('value_color', 'white')  # Default to white if no color
            if color in color_fills:
                for col_idx in range(1, len(new_adp_df.columns)):  # Exclude the DRAFTED column
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = color_fills[color]

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Add conditional formatting for drafted players
    add_conditional_formatting(worksheet, len(new_adp_df.columns))

def create_fantasypros_adp_comparison_sheet(writer, league_size=12):
    """
    Create a FantasyPros ADP comparison sheet, matching FFC ADP to FantasyPros rankings.
    Only FantasyPros columns, ADP, and value columns are included (no unified big board columns).
    Extra columns from the FantasyPros CSV are appended at the end.
    """
    # Load FantasyPros CSV
    csv_path = 'data/FantasyPros_2025_Draft_ALL_Rankings.csv'
    df_fp = pd.read_csv(csv_path)
    # Clean up column names
    df_fp.columns = [c.strip().replace('"', '').replace("'", '') for c in df_fp.columns]
    # Normalize player names for matching
    from adp_comparison import normalize_player_name, get_average_adp, get_value_color, get_value_recommendation
    df_fp['normalized_name'] = df_fp['PLAYER NAME'].apply(normalize_player_name)
    # Get FFC ADP data
    adp_df = get_average_adp(league_size)
    adp_df['normalized_name'] = adp_df['player_name'].apply(normalize_player_name)
    adp_dict = dict(zip(adp_df['normalized_name'], adp_df['adp']))
    # Build output rows
    rows = []
    for idx, row in df_fp.iterrows():
        norm_name = row['normalized_name']
        adp = adp_dict.get(norm_name, np.nan)
        # Position columns
        qb, wr, rb, te = '', '', '', ''
        pos = str(row['POS']).upper().replace('"', '').replace("'", '')
        player_name = row['PLAYER NAME']
        if 'QB' in pos:
            qb = player_name
        elif 'WR' in pos:
            wr = player_name
        elif 'RB' in pos:
            rb = player_name
        elif 'TE' in pos:
            te = player_name
        # Value metrics (rank diff, color, recommendation)
        try:
            rk = float(row['RK'])
        except Exception:
            rk = np.nan
        rank_difference = rk - adp if not np.isnan(adp) and not np.isnan(rk) else np.nan
        league_size_adjusted_diff = rank_difference / league_size if not np.isnan(rank_difference) else np.nan
        value_color = get_value_color(league_size_adjusted_diff)
        value_recommendation = get_value_recommendation(league_size_adjusted_diff)
        # Build row dict (NO unified big board columns)
        out_row = {
            'DRAFTED': 'NO',
            'ADP': adp,
            'QB': qb,
            'WR': wr,
            'RB': rb,
            'TE': te,
            'FANTASYPROS RANK': rk,
            'RANK DIFFERENCE': rank_difference,
            'VALUE RECOMMENDATION': value_recommendation,
        }
        # Add extra columns from FantasyPros at the end
        for col in df_fp.columns:
            if col not in out_row and col != 'normalized_name':
                out_row[col] = row[col]
        out_row['value_color'] = value_color
        rows.append(out_row)
    # Build DataFrame
    out_df = pd.DataFrame(rows)
    # Sort by ADP (handle NaN values)
    out_df = out_df.sort_values('ADP', na_position='last')
    # Write to Excel
    out_df.to_excel(writer, sheet_name='FANTASY PROS ADP COMPARISON', index=False)
    worksheet = writer.sheets['FANTASY PROS ADP COMPARISON']
    # Color fills (reuse existing)
    from openpyxl.styles import PatternFill
    color_fills = {
        'teal': PatternFill(start_color='00CED1', end_color='00CED1', fill_type='solid'),
        'green': PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid'),
        'light_green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
        'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'purple': PatternFill(start_color='800080', end_color='800080', fill_type='solid'),
        'black': PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    }
    # Apply color coding to rows based on value_color (skip if drafted)
    for row_idx, (_, player) in enumerate(out_df.iterrows(), start=2):
        drafted_cell = worksheet.cell(row=row_idx, column=1)
        is_drafted = drafted_cell.value == "YES"
        if not is_drafted:
            color = player.get('value_color', 'white')
            if color in color_fills:
                for col_idx in range(1, len(out_df.columns)):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = color_fills[color]
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    # Add blackout formatting
    add_conditional_formatting(worksheet, len(out_df.columns))

def add_conditional_formatting(worksheet, num_columns):
    """
    Add conditional formatting to automatically black out rows when DRAFTED = "YES"
    """
    try:
        last_row = worksheet.max_row
        if last_row > 1:  # Only apply if there's data
            from openpyxl.formatting.rule import FormulaRule
            
            # Create a more robust formula-based rule that applies to entire rows
            # Formula: =$A2="YES" (checks if DRAFTED column equals "YES")
            # This will apply to the entire row when the condition is met
            row_blackout_rule = FormulaRule(
                formula=[f'$A2="YES"'],
                stopIfTrue=True,
                fill=PatternFill(start_color='000000', end_color='000000', fill_type='solid'),
                font=Font(color='FFFFFF', bold=True)
            )
            
            # Apply to the entire data range (excluding header)
            data_range = f"A2:{chr(65 + num_columns - 1)}{last_row}"
            worksheet.conditional_formatting.add(data_range, row_blackout_rule)
            
            # Add a second rule with a different approach for better compatibility
            # This rule applies to each cell individually but checks the DRAFTED column
            for col_idx in range(1, num_columns + 1):  # Start from 1 (column A)
                col_letter = chr(64 + col_idx)  # A, B, C, etc.
                col_range = f"{col_letter}2:{col_letter}{last_row}"
                
                # Create a rule for this specific column
                col_rule = FormulaRule(
                    formula=[f'$A2="YES"'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color='000000', end_color='000000', fill_type='solid'),
                    font=Font(color='FFFFFF', bold=True)
                )
                
                worksheet.conditional_formatting.add(col_range, col_rule)
            
        print("[INFO] Conditional formatting added for drafted players (entire rows)")
        
    except Exception as e:
        print(f"[WARNING] Could not add conditional formatting: {e}")
        print("[INFO] You can manually add conditional formatting:")
        print("1. Select all data (A2:Z{last_row})")
        print("2. Home > Conditional Formatting > New Rule")
        print("3. 'Use a formula to determine which cells to format'")
        print("4. Formula: =$A2=\"YES\"")
        print("5. Format: Black background, white text")
        print("6. Apply to all sheets")

def create_adp_comparison_sheet_with_colors(writer, df, league_size=12):
    """
    Create ADP comparison sheet with color coding based on value differences.
    """
    # Get ADP comparison data
    adp_comparison_df = create_adp_comparison_sheet(df, league_size)
    
    # Select columns for the ADP comparison sheet
    adp_columns = [
        'adp', 'player_id', 'team', 'position', 'unified_rank', 
        'rank_difference', 'league_size_adjusted_diff', 'value_recommendation',
        'unified_big_board_score', 'raw_fantasy_points', 'value_color'
    ]
    
    # Only include columns that exist
    adp_export_columns = [col for col in adp_columns if col in adp_comparison_df.columns]
    adp_export_df = adp_comparison_df[adp_export_columns].copy()
    
    # Write to Excel
    adp_export_df.to_excel(writer, sheet_name='ADP_Comparison', index=False)
    
    # Get the worksheet for formatting
    worksheet = writer.sheets['ADP_Comparison']
    
    # Define color fills
    color_fills = {
        'teal': PatternFill(start_color='00CED1', end_color='00CED1', fill_type='solid'),
        'green': PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid'),
        'light_green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
        'white': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),  # Neutral - no color
        'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'purple': PatternFill(start_color='800080', end_color='800080', fill_type='solid')
    }
    
    # Apply color coding to rows based on value_color (but skip if player is drafted)
    for row_idx, (_, player) in enumerate(adp_comparison_df.iterrows(), start=2):  # Start at 2 to skip header
        # Check if this player is marked as drafted
        drafted_cell = worksheet.cell(row=row_idx, column=1)  # Column A is DRAFTED
        is_drafted = drafted_cell.value == "YES"
        
        if not is_drafted:  # Only apply value colors if not drafted
            color = player.get('value_color', 'white')  # Default to white if no color
            if color in color_fills:
                for col_idx in range(1, len(adp_export_columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = color_fills[color]
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width 